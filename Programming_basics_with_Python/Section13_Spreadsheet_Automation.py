import openpyxl 
import pprint

#open target excel file 

inv_file = openpyxl.load_workbook("inventory.xlsx") 

#retrieve sheet1 from the excel file

product_list = inv_file["Sheet1"] 

#Get the number of product per supplier#  

# (1) Starting with an empty dictionary 

products_per_supplier = {}  
total_value_per_supplier = {} 
product_under_10_inv = {}

# (2) Looping through the excel rows to fetch required values 
#       Skipping the headers and going all the way down to the last row
#       Iterate through the sheet to get all the supplier Names, Column 4
#       The iteration starts counting the supplier names and fill in the products_per_supplier dict
#       If the dict does not contain the  supplier name, else statement increment that supplier's value to 1
#       Since we have 3 suppliers, this will happen 3 times
#       Then the statement proceed to add 1 each time the supplier name matches Supplier AAA, BBB or CCC. 

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row,1).value 
    #overwriting current value of the 5th column of the inventory.xlsx
    inventory_price = product_list.cell(product_row, 5) 
    

    # Number of Products per supplier
    if supplier_name in products_per_supplier:
        current_number = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_number + 1
    else:
        print("Adding a new supplier ; ", supplier_name)
        products_per_supplier[supplier_name] = 1

    # Number of inventory per supplier 
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price 
    
    # Number of suppliers with inventory less than 10 
    if inventory < 10: 
        product_under_10_inv[int(product_number)] = inventory

    #Logic to calculate total inventory price 
    inventory_price.value = inventory * price



    
# Outputs the content of the dictionary products_per_supplier 
print(products_per_supplier) 
print(total_value_per_supplier)
print(product_under_10_inv) 


#save the values into a new excel file. 
inv_file.save("inventory_with_total_value.xlsx")






