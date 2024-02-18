import openpyxl 

#open target excel file 

inv_file = openpyxl.load_workbook("inventory.xlsx") 

#retrieve sheet1 from the excel file

product_list = inv_file["Sheet1"] 

#Get the number of product per supplier#  

# (1) Starting with an empty dictionary 

products_per_supplier = {}  

# (2) Looping through the excel rows to fetch required values 
#       Skipping the headers and going all the way down to the last row
#       Iterate through the sheet to get all the supplier Names, Column 4
#       The iteration starts counting the supplier names and fill in the products_per_supplier dict
#       If the dict does not contain the  supplier name, else statement increment that supplier's value to 1
#       Since we have 3 suppliers, this will happen 3 times
#       Then the statement proceed to add 1 each time the supplier name matches Supplier AAA, BBB or CCC. 

for product_row in range(2, product_list.max_row + 1):
   supplier_name = product_list.cell(product_row, 4).value
   #print(supplier_name)

    #Number of Products per supplier        
   if supplier_name in products_per_supplier:
       current_number = products_per_supplier[supplier_name] 
       products_per_supplier[supplier_name] = current_number + 1 

   else:
       print("Adding a new supplier")
       products_per_supplier[supplier_name] = 1 
    
# Outputs the content of the dictionary products_per_supplier 
print(products_per_supplier)





